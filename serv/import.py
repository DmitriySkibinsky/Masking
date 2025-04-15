import csv
import openpyxl

def xlsx_to_csv_with_detection(input_xlsx, output_csv):
    # Загружаем книгу Excel
    wb = openpyxl.load_workbook(input_xlsx, data_only=True)
    sheet = wb.active

    # 1. Определяем правую границу (по первой строке)
    max_col = 1
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value is None:
            break
        max_col = col

    # 2. Определяем нижнюю границу (по первому столбцу)
    max_row = 1
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value is None:
            break
        max_row = row

    # 3. Собираем данные в найденном прямоугольнике
    data = []
    for row in range(1, max_row + 1):
        row_data = []
        for col in range(1, max_col + 1):
            cell = sheet.cell(row=row, column=col)
            row_data.append(cell.value if cell.value is not None else '')
        data.append(row_data)

    # 4. Записываем в CSV
    with open(output_csv, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerows(data)

    return f"Successfully converted. Data range: A1:{openpyxl.utils.get_column_letter(max_col)}{max_row}"

# xlsx_to_csv_with_detection('soil_pollution_diseases.xlsx', 'output.csv')